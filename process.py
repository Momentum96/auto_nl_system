import re
import pandas as pd
import pymysql.cursors
from sqlalchemy import true
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import cv2
import datetime


def load_sunRS(year, month, day):
    connection = pymysql.connect(host='210.102.142.14', port=3306, user='witlab', passwd='defacto8*',
                                 charset='utf8', autocommit=True,
                                 cursorclass=pymysql.cursors.DictCursor)
    cursor = connection.cursor()
    target_date = year+'-'+month+'-'+day
    sql = "SELECT * FROM naturallight.sun_info WHERE Date='"+'%s'%target_date +"'"
    # print(sql)
    cursor.execute(sql)
    result = cursor.fetchall()
    df = pd.DataFrame(result)
    df["Sunup_Time"] =  df["Sunup_Time"].astype(str)
    df["Sunup_Time"] = df["Sunup_Time"].apply(lambda x: re.sub("0 days ", "", x))

    df["Sundown_Time"] = df["Sundown_Time"].astype(str)
    df["Sundown_Time"] = df["Sundown_Time"].apply(lambda x: re.sub("0 days ", "", x))
    # print(df)
    cursor.close()
    return df

def load_cas_data(year, month, day):
    connection = pymysql.connect(host='210.102.142.14', port=3306, user='witlab', passwd='defacto8*',
                                 charset='utf8', autocommit=True,
                                 cursorclass=pymysql.cursors.DictCursor)
    cursor = connection.cursor()
    target_date = year + '-' + month + '-' + day
    sql = "SELECT Date, CAS_SWR, CAS_MWR, CAS_LWR, CAS_446to477 FROM cas_db.cas_wave_ratio WHERE date(Date)='"+'%s'%target_date +"'"
    cursor.execute(sql)
    result = cursor.fetchall()
    cas_wave_ratio = pd.DataFrame(result)
    # print(cas_wave_ratio)

    sql = "SELECT Date, cct, photometric, color_coordinate_x, color_coordinate_y FROM cas_db.natural_tracker WHERE date(Date)='"+'%s'%target_date +"'"
    cursor.execute(sql)
    result = cursor.fetchall()
    natural_tracker = pd.DataFrame(result)
    # print(natural_tracker)

    sql = "SELECT Date, CAS_UVB,CAS_UVA, CAS_IU, CAS_EryUVI, CAS_UVB_RATIO, CAS_UVA_RATIO FROM cas_db.cas_uv_ratio WHERE date(Date)='"+'%s'%target_date +"'"
    cursor.execute(sql)
    result = cursor.fetchall()
    cas_uv_ratio = pd.DataFrame(result)
    # print(cas_uv_ratio)

    total_df = cas_wave_ratio.merge(natural_tracker, how="inner", on=None, sort=False)
    total_df = total_df.merge(cas_uv_ratio, how="inner", on=None, sort=False)
    cursor.close()

    total_df["Date"] = total_df["Date"].astype(str)
    total_df["time"] = total_df["Date"].apply(lambda x: x.split(" ")[1][0:5])

    # print(total_df)
    return total_df

def load_altitude_data(year, month, day):
    connection = pymysql.connect(host='210.102.142.14', port=3306, user='witlab', passwd='defacto8*',
                                 charset='utf8', autocommit=True,
                                 cursorclass=pymysql.cursors.DictCursor)
    cursor = connection.cursor()
    target_date = year + '-' + month + '-' + day
    sql = "SELECT Date, kasi_altitude FROM naturallight.row_altitude WHERE date(Date)='"+'%s'%target_date +"'"
    cursor.execute(sql)
    result = cursor.fetchall()
    row_altitude = pd.DataFrame(result)

    cursor.close()
    row_altitude["Date"] = row_altitude["Date"].astype(str)
    row_altitude["time"] = row_altitude["Date"].apply(lambda x: x.split(" ")[1][0:5])
    # print(row_altitude)

    return row_altitude

def load_row_weather_data(year, month, day):
    connection = pymysql.connect(host='210.102.142.14', port=3306, user='witlab', passwd='defacto8*',
                                 charset='utf8', autocommit=True,
                                 cursorclass=pymysql.cursors.DictCursor)
    cursor = connection.cursor()

    target_date = year + '-' + month + '-' + day
    sql = "SELECT Date, UVIndex, TempOut, HumOut FROM naturallight.row_weather WHERE date(Date)='"+'%s'%target_date +"'"
    cursor.execute(sql)
    result = cursor.fetchall()
    row_weather = pd.DataFrame(result)
    # print(row_weather)

    cursor.close()

    row_weather["Date"] = row_weather["Date"].astype(str)
    row_weather["time"] = row_weather["Date"].apply(lambda x: x.split(" ")[1][0:5])
    return row_weather

def select_rows(pdf: pd.DataFrame, standard_type, set_time):
    term = 0
    start = 0
    end = 0
    if standard_type == 'sunrise':
        term = 10
        start = -60
        end = +80
    elif standard_type == 'lunch':
        term = 30
        start = -180
        end = +220
    elif standard_type == 'sunset':
        term = 10
        start = -70
        end = +70

    cols = pdf.columns

    import datetime as dt
    table_list = []

    t = dt.datetime.strptime(set_time, '%H:%M')
    for i in range(start,end,term):
        temp =0;
        while abs(temp)<11:
            temp_time = (t + dt.timedelta(minutes = i+temp)).strftime('%H:%M')
            for j, row in pdf.iterrows():
                if row["time"] == temp_time:
                    table_list.append(row)
                    temp =12
            if temp ==12:
                temp_time = (t + dt.timedelta(minutes=i)).strftime('%H:%M')
                temp_row=[]
                temp_row.append(temp_time)
                for i in range(len(cols)+1):
                    temp_row.append("NULL")
            if temp==0:
                temp +=1
            elif temp>0:
                temp *= -1
            elif temp<0:
                temp = abs(temp)+1

    df = pd.DataFrame(table_list, columns=cols)
    # print(df)
    return df

def make_daily_graph(sunRS_df, Cas_df, altitude_df, row_weather_df):
    wavelength_ratio_graph(Cas_df, sunRS_df, 1)
    scatter_graph('CCT', Cas_df['time'].values, Cas_df['cct'].values, 'CCT(K)', np.arange(0, 21000, 1000), sunRS_df, 2)
    scatter_graph('Illuminance', Cas_df['time'].values, Cas_df['photometric'].values, 'Lux',
                  np.arange(0, 140000, 10000), sunRS_df, 3)
    c_xy_graph(Cas_df, sunRS_df, 4)
    scatter_graph('UV-B', Cas_df['time'].values, Cas_df['CAS_UVB'].values, 'W/m2', np.arange(0, 3.2, 0.2), sunRS_df, 5)
    scatter_line_bar_graph(Cas_df, row_weather_df, 6)
    uv_ratio_graph(Cas_df, sunRS_df, 7)
    altitude_graph(altitude_df, sunRS_df, 8)

def find_nearest(array, value):
    array = np.asarray(array)
    idx = (np.abs(array - value)).argmin()
    return idx



# 조도, 색온도, UVB 그래프에 사용
def scatter_graph(title, time, y, y_label, y_range, sunRS_df, num, y2=None, y2_name=None):
    sr_t = sunRS_df['Sunup_Time'].values[0][:-3]
    ss_t = sunRS_df['Sundown_Time'].values[0][:-3]

    start = sr_t[0] + str(int(sr_t[1]) - 1) + sr_t[2:]
    if int(ss_t[1]) == 9:
        end = str(int(ss_t[0]) + 1) + '0' + ss_t[2:]
    else:
        end = ss_t[0] + str(int(ss_t[1]) + 1) + ss_t[2:]

    time_dt = []

    for i in range(len(time)):
        time_dt.append(datetime.datetime.strptime(time[i], '%H:%M'))

    start_dt = datetime.datetime.strptime(start, '%H:%M')
    srt_dt = datetime.datetime.strptime(sr_t, '%H:%M')
    sst_dt = datetime.datetime.strptime(ss_t, '%H:%M')
    end_dt = datetime.datetime.strptime(end, '%H:%M')

    start_index = find_nearest(time_dt, start_dt)
    srt_index = find_nearest(time_dt, srt_dt)
    sst_index = find_nearest(time_dt, sst_dt)
    end_index = find_nearest(time_dt, end_dt)

    x = []
    for i in range(len(time)):
        x.append(datetime.datetime.strptime(time[i], '%H:%M'))

    plt.figure(figsize=(9, 8))

    plt.scatter(x[start_index:srt_index], y[start_index:srt_index], c='g', s=5)
    plt.scatter(x[srt_index:sst_index], y[srt_index:sst_index], c='r', s=5)
    plt.scatter(x[sst_index:end_index], y[sst_index:end_index], c='g', s=5)

    plt.gcf().autofmt_xdate()
    myFmt = mdates.DateFormatter('%H:%M')
    plt.gca().xaxis.set_major_formatter(myFmt)
    plt.grid()
    plt.ylim([0, y_range[-1]])
    plt.yticks(y_range)
    plt.xlabel('Time', size=16, fontweight='bold')
    plt.ylabel(y_label, size=16, fontweight='bold')
    plt.title(title, size=24, fontweight='bold')
    # plt.show()
    plt.savefig(str(num) + '.png')

# 색좌표 그래프
def c_xy_graph(Cas_df, sunRS_df, num, y2=None, y2_name=None):
    time = Cas_df['time'].values
    x = Cas_df['color_coordinate_x'].values
    y = Cas_df['color_coordinate_y'].values

    x = (x - 0.2) / (0.55 - 0.2) * 484
    y = (y - 0.2) / (0.5 - 0.2) * 414

    sr_t = sunRS_df['Sunup_Time'].values[0][:-3]
    ss_t = sunRS_df['Sundown_Time'].values[0][:-3]

    start = sr_t[0] + str(int(sr_t[1]) - 1) + sr_t[2:]
    if int(ss_t[1]) == 9:
        end = str(int(ss_t[0]) + 1) + '0' + ss_t[2:]
    else:
        end = ss_t[0] + str(int(ss_t[1]) + 1) + ss_t[2:]

    time_dt = []

    for i in range(len(time)):
        time_dt.append(datetime.datetime.strptime(time[i], '%H:%M'))

    start_dt = datetime.datetime.strptime(start, '%H:%M')
    srt_dt = datetime.datetime.strptime(sr_t, '%H:%M')
    sst_dt = datetime.datetime.strptime(ss_t, '%H:%M')
    end_dt = datetime.datetime.strptime(end, '%H:%M')

    start_index = find_nearest(time_dt, start_dt)
    srt_index = find_nearest(time_dt, srt_dt)
    sst_index = find_nearest(time_dt, sst_dt)
    end_index = find_nearest(time_dt, end_dt)

    image = cv2.imread("c_xy.jpg")
    image = cv2.flip(image, 0)

    plt.figure(figsize=(9, 8))
    plt.imshow(image)

    plt.scatter(x[start_index:srt_index], y[start_index:srt_index], c='g', s=20, marker='v')
    plt.scatter(x[srt_index:sst_index], y[srt_index:sst_index], c='r', s=20, marker='v')
    plt.scatter(x[sst_index:end_index], y[sst_index:end_index], c='g', s=20, marker='v')

    plt.grid()
    x_t = []
    for i in range(20, 60, 5):
        x_t.append(i / 100)
    plt.xticks((np.arange(0.20, 0.60, 0.05) - 0.2) / (0.55 - 0.2) * 484, x_t)
    plt.xlim([0, 484])
    y_t = []
    for i in range(20, 52, 2):
        y_t.append(i / 100)
    plt.yticks((np.arange(0.20, 0.52, 0.02) - 0.2) / (0.5 - 0.2) * 414, y_t)
    plt.ylim([0, 414])
    plt.xlabel('x', size=16, fontweight='bold')
    plt.ylabel('y', size=16, fontweight='bold')
    plt.title('Chromaticity Coordinates', size=24, fontweight='bold')
    # plt.gca().invert_yaxis()

    # plt.show()
    plt.savefig(str(num) + '.png')

def scatter_line_bar_graph(Cas_df, row_weather_df, num):
    time = Cas_df['time'].values
    w_time = row_weather_df['time'].values

    # print(w_time)

    x_range = ['07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30',
               '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00',
               '18:30']
    x_index = []

    time_dt = []
    w_time_dt = []

    for i in range(len(time)):
        time_dt.append(datetime.datetime.strptime(time[i], '%H:%M'))

    for i in range(len(w_time)):
        w_time_dt.append(datetime.datetime.strptime(w_time[i], '%H:%M'))

    x_range_dt = []

    for i in range(len(x_range)):
        x_range_dt.append(datetime.datetime.strptime(x_range[i], '%H:%M'))

    for i in range(len(x_range)):
        # x_index.append(np.where(time == x_range[i])[0][0])
        x_index.append(find_nearest(time_dt, x_range_dt[i]))

    w_x_index = []
    for i in range(len(x_range)):
        # w_x_index.append(np.where(w_time == x_range[i])[0][0])
        w_x_index.append(find_nearest(w_time_dt, x_range_dt[i]))

    x = []
    for i in range(len(time)):
        x.append(datetime.datetime.strptime(time[i], '%H:%M'))

    w_x = []
    for i in range(len(w_time)):
        w_x.append(datetime.datetime.strptime(w_time[i], '%H:%M'))

    x_data = []
    CAS_UVI = []
    IU = []
    w_x_data = []
    Weather_UVI = []

    c_uvi = Cas_df['CAS_EryUVI'].values
    c_iu = Cas_df['CAS_IU'].values

    w_uvi = row_weather_df['UVIndex'].values

    TARGET_IU = 400
    total_IU = 0

    for i in range(len(x_index) - 1):
        x_data.append(x[x_index[i]])
        CAS_UVI.append(c_uvi[x_index[i]])
        IU.append(np.sum(c_iu[x_index[i]:x_index[i + 1]]))
        total_IU += np.sum(c_iu[x_index[i]:x_index[i + 1]])
        w_x_data.append(w_x[w_x_index[i]])
        if w_uvi[w_x_index[i]]=="---":
            Weather_UVI.append(float(0))
        else:
            Weather_UVI.append(float(w_uvi[w_x_index[i]]))

    total_IU = int(total_IU)

    fig, ax1 = plt.subplots(figsize=(9, 8))
    p1 = ax1.bar(x_data, IU, width=0.01, alpha=0.6, color='r')
    ax1.set_ylabel('IU', fontweight='bold', size=16)
    ax1.set_yticks(np.arange(0, 1300, 100))
    ax1.set_ylim([0, 1200])
    ax1.text(datetime.datetime(1900, 1, 1, 9, 45), 800,
             'Target IU : ' + str(TARGET_IU) + 'IU   Total IU : ' + str(total_IU) + 'IU',
             fontsize=14, bbox={'facecolor': 'red', 'alpha': 0.5, 'pad': 5}, fontweight='bold')

    ax2 = ax1.twinx()
    p2 = ax2.plot(x_data, CAS_UVI, 'o-', c='g', label='CAS_EryUVI')
    p3 = ax2.plot(x_data, Weather_UVI, 'o-', c='b', label='Weather_UVI')
    plt.grid()
    ax2.set_xlabel('Time', fontweight='bold', size=16)
    ax2.set_ylabel('UVI', fontweight='bold', size=16)
    ax2.set_title('IU&UVI', fontweight='bold', size=24)
    ax2.set_yticks(np.arange(0, 15, 1))
    # ax1.set_ylim([0, 14])
    plt.gcf().autofmt_xdate()
    myFmt = mdates.DateFormatter('%H:%M')
    plt.gca().xaxis.set_major_formatter(myFmt)
    plt.legend((p1[0], p2[0], p3[0]), ('IU', 'CAS_EryUVI', 'Weather_UVI'), loc='upper center')
    # plt.show()
    plt.savefig(str(num) + '.png')

# 파장비율 그래프
def wavelength_ratio_graph(Cas_df, sunRS_df, num):
    time = Cas_df['time'].values

    sr_t = sunRS_df['Sunup_Time'].values[0][:-3]
    ss_t = sunRS_df['Sundown_Time'].values[0][:-3]

    start = sr_t[0] + str(int(sr_t[1]) - 1) + sr_t[2:]
    if int(ss_t[1]) == 9:
        end = str(int(ss_t[0])+1) + '0' + ss_t[2:]
    else:
        end = ss_t[0] + str(int(ss_t[1]) + 1) + ss_t[2:]

    time_dt = []

    for i in range(len(time)):
        time_dt.append(datetime.datetime.strptime(time[i], '%H:%M'))

    start_dt = datetime.datetime.strptime(start, '%H:%M')
    end_dt = datetime.datetime.strptime(end, '%H:%M')

    start_index = find_nearest(time_dt, start_dt)
    end_index = find_nearest(time_dt, end_dt)

    x = []
    for i in range(len(time)):
        x.append(datetime.datetime.strptime(time[i], '%H:%M'))

    swr = Cas_df['CAS_SWR'].values
    mwr = Cas_df['CAS_MWR'].values
    lwr = Cas_df['CAS_LWR'].values

    plt.figure(figsize=(9, 8))
    labels = ['swr', 'mwr', 'lwr']
    plt.stackplot(x[start_index:end_index], swr[start_index:end_index], mwr[start_index:end_index],
                  lwr[start_index:end_index], colors=['royalblue', 'limegreen', 'coral'], labels=labels)

    plt.gcf().autofmt_xdate()
    myFmt = mdates.DateFormatter('%H:%M')
    plt.gca().xaxis.set_major_formatter(myFmt)
    plt.grid()
    plt.ylim([0, 100])
    plt.yticks(np.arange(0, 110, 10))
    plt.xlabel('Time', size=16, fontweight='bold')
    plt.ylabel('Ratio (%)', size=16, fontweight='bold')
    plt.title('Wavelength Ratio', size=24, fontweight='bold')
    plt.legend(loc='upper center', fontsize='large')
    # plt.show()
    plt.savefig(str(num) + '.png')

def uv_ratio_graph(Cas_df, sunRS_df, num):
    time = Cas_df['time'].values

    sr_t = sunRS_df['Sunup_Time'].values[0][:-3]
    ss_t = sunRS_df['Sundown_Time'].values[0][:-3]

    start = sr_t[0] + str(int(sr_t[1]) - 1) + sr_t[2:]
    if int(ss_t[1]) == 9:
        end = str(int(ss_t[0]) + 1) + '0' + ss_t[2:]
    else:
        end = ss_t[0] + str(int(ss_t[1]) + 1) + ss_t[2:]

    time_dt = []

    for i in range(len(time)):
        time_dt.append(datetime.datetime.strptime(time[i], '%H:%M'))

    start_dt = datetime.datetime.strptime(start, '%H:%M')
    end_dt = datetime.datetime.strptime(end, '%H:%M')

    start_index = find_nearest(time_dt, start_dt)
    end_index = find_nearest(time_dt, end_dt)

    x = []
    for i in range(len(time)):
        x.append(datetime.datetime.strptime(time[i], '%H:%M'))

    UVA = Cas_df['CAS_UVA_RATIO'].values
    UVB = Cas_df['CAS_UVB_RATIO'].values

    plt.figure(figsize=(9, 8))
    labels = ['CAS_UVB_RATIO', 'CAS_UVA_RATIO']
    plt.stackplot(x[start_index:end_index], UVB[start_index:end_index], UVA[start_index:end_index],
                  colors=['royalblue', 'limegreen'], labels=labels)

    plt.gcf().autofmt_xdate()
    myFmt = mdates.DateFormatter('%H:%M')
    plt.gca().xaxis.set_major_formatter(myFmt)
    plt.grid()
    plt.ylim([0, 20])
    plt.yticks(np.arange(0, 21, 1))
    plt.xlabel('Time', size=16, fontweight='bold')
    plt.ylabel('Ratio (%)', size=16, fontweight='bold')
    plt.title('UV-A&UV-B Ratio', size=24, fontweight='bold')
    plt.legend(loc='upper center', fontsize='large')
    # plt.show()
    plt.savefig(str(num) + '.png')

def altitude_graph(altitude_df, sunRS_df, num):
    time = altitude_df['time'].values

    sr_t = sunRS_df['Sunup_Time'].values[0][:-3]
    ss_t = sunRS_df['Sundown_Time'].values[0][:-3]

    start = sr_t[0] + str(int(sr_t[1]) - 1) + sr_t[2:]
    if int(ss_t[1]) == 9:
        end = str(int(ss_t[0]) + 1) + '0' + ss_t[2:]
    else:
        end = ss_t[0] + str(int(ss_t[1]) + 1) + ss_t[2:]

    start_index = 0
    end_index = 0

    for i in range(len(time)):
        if time[i] >= start:
            start_index = i
            break

    for i in range(len(time)):
        if time[i] >= end:
            end_index = i
            break

    x = []
    for i in range(len(time)):
        x.append(datetime.datetime.strptime(time[i], '%H:%M'))

    y = altitude_df['kasi_altitude'].values

    plt.figure(figsize=(9, 8))

    plt.plot(x[start_index:end_index], y[start_index:end_index], c='r', label='kasi_altitude')

    plt.gcf().autofmt_xdate()
    myFmt = mdates.DateFormatter('%H:%M')
    plt.gca().xaxis.set_major_formatter(myFmt)
    plt.grid()
    plt.ylim([0, 100])
    plt.yticks(np.arange(0, 105, 5))
    plt.xlabel('Time', size=16, fontweight='bold')
    plt.ylabel('ft', size=16, fontweight='bold')
    plt.title('Altitude', size=24, fontweight='bold')
    plt.legend(loc='upper center', fontsize='large')
    # plt.show()
    plt.savefig(str(num) + '.png')





def write_to_excel(sunRS_df, tab_1_1, tab_1_2 , tab_1_3 , tab_2_1, tab_2_2, tab_2_3 ):
    import openpyxl
    import clipboard as clp

    wb = openpyxl.Workbook()
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])

##두번째 시트먼저 해야함.
    ws1 = wb.create_sheet(title='naturallight_uv', index=0)

    ws1.merge_cells('A1:H20')
    my_png = openpyxl.drawing.image.Image('5.png')
    my_png.height = 439.55905512
    my_png.width = 576
    ws1.add_image(my_png, 'A1')

    ws1.merge_cells('I1:P20')
    my_png = openpyxl.drawing.image.Image('6.png')
    my_png.height = 439.55905512
    my_png.width = 576
    ws1.add_image(my_png, 'I1')

    ws1.merge_cells('A23:H42')
    my_png = openpyxl.drawing.image.Image('7.png')
    my_png.height = 439.55905512
    my_png.width = 576
    ws1.add_image(my_png, 'A23')

    ws1.merge_cells('I23:P42')
    my_png = openpyxl.drawing.image.Image('8.png')
    my_png.height = 439.55905512
    my_png.width = 576
    ws1.add_image(my_png, 'I23')

    ws1.merge_cells('A21:H22')
    ws1['A21'] = "CAS_UVB  :  CAS          (280-315nm)"
    ws1['A21'].font =openpyxl.styles.Font(size=9, name ='Calibri')
    ws1['A21'].alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')


    ws1.merge_cells('I21:P22')
    ws1['I21'] = "IU  :  CAS                        (280-400nm)\nCAS_EryUVI  :  CAS              (280-400nm)\nWeather_UVI : 기상관측기   (280-360nm)"
    ws1['I21'].font =openpyxl.styles.Font(size=9, name ='Calibri')
    ws1['I21'].alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')


    ws1.merge_cells('A43:H44')
    ws1['A43'] = "CAS_UVB_RATIO : CAS (280-315nm)\nCAS_UVA_RATIO : CAS (315-400nm)"
    ws1['A43'].font =openpyxl.styles.Font(size=9, name ='Calibri')
    ws1['A43'].alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')

    ws1.merge_cells('I43:P44')
    ws1['I43'] = "kasi_altitude : 천문학사이트_고도각"
    ws1['I43'].font =openpyxl.styles.Font(size=9, name ='Calibri')
    ws1['I43'].alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')


    ##박스
    thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                 right=openpyxl.styles.borders.Side(style='thin'),
                                                 top=openpyxl.styles.borders.Side(style='thin'),
                                                 bottom=openpyxl.styles.borders.Side(style='thin'))
    #
    # Copy dataframe to clipboard
    tab_2_1.to_clipboard()
    # paste the clipboard to a valirable
    cells = clp.paste()
    # split text in varialble as rows and columns
    cells = [x.split() for x in cells.split('\n')]
    #
    # Get the Sheet
    sheet = ws1
    # Paste clipboard values to the sheet
    for i, r in zip(range(46, len(cells) + 1 + 46), cells):
        for j, c in zip(range(2, len(r) + 2), r):
            if len(c)>7:
                c = c[0:6]
            sheet.cell(row=i, column=j).value = c
            sheet.cell(row=i, column=j).border = thin_border
            sheet.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    #

    #
    # Copy dataframe to clipboard
    tab_2_2.to_clipboard()
    # paste the clipboard to a valirable
    cells = clp.paste()
    # split text in varialble as rows and columns
    cells = [x.split() for x in cells.split('\n')]
    #
    # Get the Sheet
    sheet = ws1
    # Paste clipboard values to the sheet
    for i, r in zip(range(58, len(cells) + 1 + 58), cells):
        for j, c in zip(range(2, len(r) + 2), r):
            if len(c) > 7:
                c = c[0:6]
            sheet.cell(row=i, column=j).value = c
            sheet.cell(row=i, column=j).border = thin_border
            sheet.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    #

    #
    # Copy dataframe to clipboard
    tab_2_3.to_clipboard()
    # paste the clipboard to a valirable
    cells = clp.paste()
    # split text in varialble as rows and columns
    cells = [x.split() for x in cells.split('\n')]
    #
    # Get the Sheet
    sheet = ws1
    # Paste clipboard values to the sheet
    for i, r in zip(range(70, len(cells) + 1 + 70), cells):
        for j, c in zip(range(2, len(r) + 2), r):
            if len(c) > 7:
                c = c[0:6]
            sheet.cell(row=i, column=j).value = c
            sheet.cell(row=i, column=j).border = thin_border
            sheet.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    #

    row_name = ["상대 시간", "절대 시간", "UV-A W/m^2", "UV-B W/m^2", "UV-A %", "UV-B %", "비타민 D IU", "CAS UVI", "기상관측기 UVI", "기상관측기 온도(ºC)", "기상관측기 습도(%)"]
    for i in range(46, 57):
        ws1.merge_cells('A%d:B%d' % (i, i))
        ws1['A%d' % i] = row_name[i - 46]
        ws1['A%d' % i].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws1['A%d' % i].border = thin_border
        ws1['B%d' % i].border = thin_border
        ws1['I%d' % i].fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                               fgColor=openpyxl.styles.colors.Color(rgb='FFCC00'))
        ws1['P%d' % i].font = openpyxl.styles.Font(color='3366FF', bold=True)
    for i in range(58, 69):
        ws1.merge_cells('A%d:B%d' % (i, i))
        ws1['A%d' % i] = row_name[i - 58]
        ws1['A%d' % i].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws1['A%d' % i].border = thin_border
        ws1['B%d' % i].border = thin_border
        ws1['I%d' % i].fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                               fgColor=openpyxl.styles.colors.Color(rgb='FFCC00'))
        ws1['C%d' % i].font = openpyxl.styles.Font(color='800000', bold=True)
    for i in range(70, 81):
        ws1.merge_cells('A%d:B%d' % (i, i))
        ws1['A%d' % i] = row_name[i - 70]
        ws1['A%d' % i].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws1['A%d' % i].border = thin_border
        ws1['B%d' % i].border = thin_border
        ws1['J%d' % i].fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                               fgColor=openpyxl.styles.colors.Color(rgb='FFCC00'))
        ws1['C%d' % i].font = openpyxl.styles.Font(color='3366FF', bold=True)

    ws1['A58'] = "절대 시간"
    col_name_1 = ["일출-60", "일출-50", "일출-40", "일출-30", "일출-20", "일출-10", "일출", "일출+10", "일출+20", "일출+30", "일출+40",
                  "일출+50", "일출+60", "일출+70"]
    col_name_2 = ["9:00", "9:30", "10:00", "10:30", "11:00", "11:30", "12:00", "12:30", "13:00", "13:30", "14:00",
                  "14:30", "15:00", "15:30"]
    col_name_3 = ["일몰-70", "일몰-60", "일몰-50", "일몰-40", "일몰-30", "일몰-20", "일몰-10", "일몰", "일몰+10", "일몰+20", "일몰+30",
                  "일몰+40", "일몰+50", "일몰+60"]

    for i in range(len(col_name_1)):
        ws1.cell(row=46, column=i + 3).value = col_name_1[i]
        ws1.cell(row=46, column=i + 3).border = thin_border
        ws1.cell(row=46, column=i + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws1.cell(row=58, column=i + 3).value = col_name_2[i]
        ws1.cell(row=58, column=i + 3).border = thin_border
        ws1.cell(row=58, column=i + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws1.cell(row=59, column=i + 3).value = col_name_2[i]
        ws1.cell(row=59, column=i + 3).border = thin_border
        ws1.cell(row=59, column=i + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws1.cell(row=70, column=i + 3).value = col_name_3[i]
        ws1.cell(row=70, column=i + 3).border = thin_border
        ws1.cell(row=70, column=i + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('K81:M81')
    ws1['K81'] = "일출, 정오, 일몰"
    ws1['K81'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws1['K81'].border = thin_border
    ws1['L81'].border = thin_border
    ws1['M81'].border = thin_border

    ws1.merge_cells('N81:P81')
    ws1['N81'] = "일출 후, 일몰 전 최소"
    ws1['N81'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws1['N81'].border = thin_border
    ws1['O81'].border = thin_border
    ws1['P81'].border = thin_border

    ws1['K81'].fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                       fgColor=openpyxl.styles.colors.Color(rgb='FFCC00'))
    ws1['N81'].font = openpyxl.styles.Font(color='3366FF', bold=True)

    ws1.merge_cells('A82:P85')
    ws1['A82'].alignment = openpyxl.styles.Alignment(wrapText=True)
    ws1['A82'] = "    작성자  : \n    관측자  : \n    특이사항 :"
    ws1['A82'].font = openpyxl.styles.Font(size=14, name='Calibri')

    ########################


    ws = wb.create_sheet(title='naturallight', index=0)

    ws.merge_cells('A1:P2')
    ws['A1'] = str(sunRS_df["Date"][0])+" (Sunrise : "+ sunRS_df["Sunup_Time"][0][:5]+", Sunset : "+ sunRS_df["Sundown_Time"][0][:5] +")"
    ws['A1'].font =openpyxl.styles.Font(size=20, name ='Calibri')
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:H22')
    ws['A3'] = "조도 그래프"
    ws['A3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    my_png = openpyxl.drawing.image.Image('1.png')
    my_png.height = 439.55905512
    my_png.width = 576
    ws.add_image(my_png, 'A3')

    ws.merge_cells('I3:P22')
    ws['I3'] = "CCT 그래프"
    ws['I3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    my_png = openpyxl.drawing.image.Image('2.png')
    my_png.height = 439.55905512
    my_png.width = 576
    ws.add_image(my_png, 'I3')

    ws.merge_cells('A23:H42')
    ws['A23'] = "SWR 그래프"
    ws['A23'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    my_png = openpyxl.drawing.image.Image('3.png')
    my_png.height = 439.55905512
    my_png.width = 576
    ws.add_image(my_png, 'A23')

    ws.merge_cells('I23:P42')
    ws['I23'] = "Chromatictiy 그래프"
    ws['I23'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    my_png = openpyxl.drawing.image.Image('4.png')
    my_png.height = 439.55905512
    my_png.width = 576
    ws.add_image(my_png, 'I23')

    ##박스
    thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                 right=openpyxl.styles.borders.Side(style='thin'),
                                                 top=openpyxl.styles.borders.Side(style='thin'),
                                                 bottom=openpyxl.styles.borders.Side(style='thin'))
    # my_style = openpyxl.styles(border=thin_border)
    # ws['A44'].border = thin_border
    #

    #
    # Copy dataframe to clipboard
    tab_1_1.to_clipboard()
    # paste the clipboard to a valirable
    cells = clp.paste()
    # split text in varialble as rows and columns
    cells = [x.split() for x in cells.split('\n')]
    #
    # Get the Sheet
    sheet = ws
    # Paste clipboard values to the sheet
    for i, r in zip(range(44, len(cells)+1+44), cells):
        for j, c in zip(range(2, len(r)+2), r):
            if len(c) > 7:
                c = c[0:6]
            sheet.cell(row=i, column=j).value = c
            sheet.cell(row=i, column=j).border = thin_border
            sheet.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    #

    #
    # Copy dataframe to clipboard
    tab_1_2.to_clipboard()
    # paste the clipboard to a valirable
    cells = clp.paste()
    # split text in varialble as rows and columns
    cells = [x.split() for x in cells.split('\n')]
    #
    # Get the Sheet
    sheet = ws
    # Paste clipboard values to the sheet
    for i, r in zip(range(55, len(cells) + 1 + 55), cells):
        for j, c in zip(range(2, len(r) + 2), r):
            if len(c) > 7:
                c = c[0:6]
            sheet.cell(row=i, column=j).value = c
            sheet.cell(row=i, column=j).border = thin_border
            sheet.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    #

    #
    # Copy dataframe to clipboard
    tab_1_3.to_clipboard()
    # paste the clipboard to a valirable
    cells = clp.paste()
    # split text in varialble as rows and columns
    cells = [x.split() for x in cells.split('\n')]
    #
    # Get the Sheet
    sheet = ws
    # Paste clipboard values to the sheet
    for i, r in zip(range(66, len(cells) + 1 + 66), cells):
        for j, c in zip(range(2, len(r) + 2), r):
            if len(c) > 7:
                c = c[0:6]
            sheet.cell(row=i, column=j).value = c
            sheet.cell(row=i, column=j).border = thin_border
            sheet.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    #

    row_name =["상대 시간", "절대 시간", "446~477", "단파장 %", "중파장 %", "장파장 %", "조도 lx", "색온도 K", "색좌표 x", "색좌표 y"]
    for i in range(44,54):
        ws.merge_cells('A%d:B%d'%(i,i))
        ws['A%d'%i] = row_name[i-44]
        ws['A%d'%i].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws['A%d' % i].border = thin_border
        ws['B%d' % i].border = thin_border
        ws['I%d' % i].fill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=openpyxl.styles.colors.Color(rgb='FFCC00'))
        ws['P%d' % i].font = openpyxl.styles.Font(color='3366FF', bold=True)
    for i in range(55, 65):
        ws.merge_cells('A%d:B%d'%(i,i))
        ws['A%d' % i] = row_name[i - 55]
        ws['A%d' % i].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws['A%d' % i].border = thin_border
        ws['B%d' % i].border = thin_border
        ws['I%d' % i].fill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=openpyxl.styles.colors.Color(rgb='FFCC00'))
        ws['C%d' % i].font = openpyxl.styles.Font(color='800000', bold=True)
    for i in range(66, 76):
        ws.merge_cells('A%d:B%d'%(i,i))
        ws['A%d' % i] = row_name[i - 66]
        ws['A%d' % i].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws['A%d' % i].border = thin_border
        ws['B%d' % i].border = thin_border
        ws['J%d' % i].fill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=openpyxl.styles.colors.Color(rgb='FFCC00'))
        ws['C%d' % i].font = openpyxl.styles.Font(color='3366FF', bold=True)

    ws['A55'] = "절대 시간"
    col_name_1 = ["일출-60", "일출-50", "일출-40", "일출-30", "일출-20", "일출-10", "일출", "일출+10", "일출+20", "일출+30", "일출+40", "일출+50", "일출+60", "일출+70"]
    col_name_2 = ["9:00", "9:30", "10:00", "10:30", "11:00", "11:30", "12:00", "12:30", "13:00", "13:30", "14:00", "14:30" ,"15:00", "15:30"]
    col_name_3 = ["일몰-70", "일몰-60", "일몰-50", "일몰-40", "일몰-30", "일몰-20", "일몰-10", "일몰", "일몰+10", "일몰+20", "일몰+30", "일몰+40", "일몰+50", "일몰+60"]

    for i in range(len(col_name_1)):
        ws.cell(row=44, column=i+3).value = col_name_1[i]
        ws.cell(row=44, column=i + 3).border = thin_border
        ws.cell(row=44, column=i + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws.cell(row=55, column=i + 3).value = col_name_2[i]
        ws.cell(row=55, column=i + 3).border = thin_border
        ws.cell(row=55, column=i + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws.cell(row=56, column=i + 3).value = col_name_2[i]
        ws.cell(row=56, column=i + 3).border = thin_border
        ws.cell(row=56, column=i + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws.cell(row=66, column=i + 3).value = col_name_3[i]
        ws.cell(row=66, column=i + 3).border = thin_border
        ws.cell(row=66, column=i + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    ws.merge_cells('K76:M76')
    ws['K76'] = "일출, 정오, 일몰"
    ws['K76'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws['K76'].border = thin_border
    ws['L76'].border = thin_border
    ws['M76'].border = thin_border

    ws.merge_cells('N76:P76')
    ws['N76'] = "일출 후, 일몰 전 최소"
    ws['N76'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws['N76'].border = thin_border
    ws['O76'].border = thin_border
    ws['P76'].border = thin_border

    ws['K76'].fill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=openpyxl.styles.colors.Color(rgb='FFCC00'))
    ws['N76'].font = openpyxl.styles.Font(color='3366FF', bold=True)

    ws.merge_cells('A77:P80')
    ws['A77'].alignment = openpyxl.styles.Alignment(wrapText=True)
    ws['A77'] = "    작성자  : \n    관측자  : \n    특이사항 :"
    ws['A77'].font =openpyxl.styles.Font(size=14, name ='Calibri')

    date_temp = str(sunRS_df["Date"][0]).split("-")
    file_name = 'natural_'+date_temp[0]+date_temp[1]+date_temp[2]+'.xlsx'
    wb.save("C:\\Users\\user\\Desktop\\nl\\"+file_name)
    wb.close()


def process(year, month, day):
    ## db별 df로드 ##
    sunRS_df = load_sunRS(year, month, day)
    Cas_df = load_cas_data(year, month, day)
    altitude_df = load_altitude_data(year, month, day)
    row_weather_df = load_row_weather_data(year, month, day)

    ## 그래프 그리기 ##
    make_daily_graph(sunRS_df, Cas_df, altitude_df, row_weather_df)

    ## 탭별 df 제작 ##
    naturallight_df = Cas_df[["time", "CAS_446to477", "CAS_SWR", "CAS_MWR", "CAS_LWR", "photometric", "cct", "color_coordinate_x", "color_coordinate_y"]]

    uv_df = Cas_df[["time", "CAS_UVA", "CAS_UVB", "CAS_UVA_RATIO", "CAS_UVB_RATIO", "CAS_IU", "CAS_EryUVI"]]
    weather_df = row_weather_df[["time", "UVIndex", "TempOut", "HumOut"]]
    nl_uv_df = pd.merge(uv_df,weather_df, how="left", on=None, sort=False)

    ## 일출, 정오, 일몰 시간 기준 일부 데이터 추출
    tab_1_1 = select_rows(naturallight_df, "sunrise", sunRS_df["Sunup_Time"][0][:5])
    tab_1_1 = tab_1_1.T
    tab_1_2 = select_rows(naturallight_df, "lunch", "12:00")
    tab_1_2 = tab_1_2.T
    tab_1_3 = select_rows(naturallight_df, "sunset", sunRS_df["Sundown_Time"][0][:5])
    tab_1_3 = tab_1_3.T

    tab_2_1 = select_rows(nl_uv_df, "sunrise", sunRS_df["Sunup_Time"][0][:5])
    tab_2_1 = tab_2_1.T
    tab_2_2 = select_rows(nl_uv_df, "lunch", "12:00")
    tab_2_2 = tab_2_2.T
    tab_2_3 = select_rows(nl_uv_df, "sunset", sunRS_df["Sundown_Time"][0][:5])
    tab_2_3 = tab_2_3.T

    # # excel로 작성하기.
    write_to_excel(sunRS_df, tab_1_1, tab_1_2 , tab_1_3 , tab_2_1, tab_2_2, tab_2_3)


def load_date_list(start_date, end_date):
    connection = pymysql.connect(host='210.102.142.14', port=3306, user='witlab', passwd='defacto8*',
                                 charset='utf8', autocommit=True,
                                 cursorclass=pymysql.cursors.DictCursor)
    cursor = connection.cursor()

    sql = "SELECT distinct(date(Date))as dates FROM cas_db.cas_wave_ratio WHERE date(Date)>='"+'%s'%start_date +"'and date(Date)<='"+'%s'%end_date +"'"
    cursor.execute(sql)
    result = cursor.fetchall()
    date_df1 = pd.DataFrame(result)

    sql = "SELECT distinct(date(Date))as dates FROM naturallight.row_altitude WHERE date(Date)>='"+'%s'%start_date +"'and date(Date)<='"+'%s'%end_date +"'"
    cursor.execute(sql)
    result = cursor.fetchall()
    date_df2 = pd.DataFrame(result)

    sql = "SELECT distinct(date(Date))as dates FROM naturallight.row_weather WHERE date(Date)>='"+'%s'%start_date +"'and date(Date)<='"+'%s'%end_date +"'"
    cursor.execute(sql)
    result = cursor.fetchall()
    date_df3 = pd.DataFrame(result)

    date_df = date_df1.merge(date_df2, how="inner", on=None, sort=False)
    date_df = date_df.merge(date_df3, how="inner", on=None, sort=False)

    date_df["dates"] = date_df["dates"].astype(str)
    cursor.close()

    date_list = date_df.values.tolist()
    # print(date_list)
    return date_list

if __name__ == '__main__':
    print("날짜 구간별-자연광정리 자동화 툴")

    print("분석 시작날짜 입력")
    s_year = input("year(yyyy) :")
    s_month = input("month(mm) :")
    s_day = input("day(dd) :")
    start_date = s_year+"-"+s_month+"-"+s_day

    print("분석 종료날짜 입력")
    e_year = input("year(yyyy) :")
    e_month = input("month(mm) :")
    e_day = input("day(dd) :")
    end_date = e_year + "-" + e_month + "-" + e_day

    date_list = load_date_list(start_date, end_date)
    for i in date_list:
        year = str(i).split("-")[0][-4:]
        month = str(i).split("-")[1]
        day = str(i).split("-")[2][:-2]
        target_date = year + '-' + month + '-' + day
        process(year, month, day)
        print(target_date+": 완료")

    ## 구간 입력후 에러가 난다면, cas/기상관측기/천문연구원 자료가 DB상에 겹치는게 단 1개도 없는 날짜구간이기 때문.